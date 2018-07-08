# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from collections import OrderedDict
import operator
import csv

from models.patent_record import Patent
from utils import extractMedicines, extractFormOfPatent, extractFunctionFromPatent, extractMajorFunctionFromPatent


def data_process(wb, sheet_name):
    if wb is None:
        print("Wb object should not be none!")
    if sheet_name is None or sheet_name == "":
        print("Sheet name should not be none!")
        return
    # Get sheet based on the sheet name
    sheet = wb[sheet_name]
    row_count = sheet.max_row

    if row_count < 2:
        print("sheet has no data!")
        return

    total_medicines = []
    total_forms = []
    total_major_functions = []
    total_functions = []

    total_medicines_set = set()
    total_forms_set = set()
    total_major_functions_set = set()
    total_functions_set = set()

    patent_medicines = []

    for i in range(2, row_count+1):

        # patent name (should not be none!)
        patent_name = ""
        if sheet["A" + str(i)].value:
            patent_name = sheet["A" + str(i)].value.strip()
            patent_name = patent_name.replace("\n", "")
        else:
            continue

        # medicine components
        medicine_components = ""
        medicines = []
        if sheet["B"+str(i)].value:
            medicine_components = sheet["B"+str(i)].value.strip().replace("\n", "").replace(" ", "")
            medicines = extractMedicines(medicine_components)
            if len(medicines) > 0:
                patent_medicines.append(medicines)

        # patent form
        form_str = ""
        forms = []
        if sheet['C' + str(i)].value:
            form_str = sheet['C' + str(i)].value.strip().replace("\n", "").replace(" ", "")
            forms = extractFormOfPatent(form_str)

        # patent major function
        major_function_str = ''
        major_functions = []
        if sheet['D' + str(i)].value:
            major_function_str = sheet['D' + str(i)].value.strip().replace("\n", "").replace(" ", "")
            major_functions = extractMajorFunctionFromPatent(major_function_str)

        # patent functions
        function_str = ''
        functions = []
        if sheet['E' + str(i)].value:
            function_str = sheet['E' + str(i)].value.strip().replace("\n", "").replace(" ", "")
            functions = extractFunctionFromPatent(function_str)

        # patent id
        patent_id = ''
        if sheet['F' + str(i)].value:
            patent_id = str(sheet['F' + str(i)].value).strip().replace("\n", "").replace(" ", "")

        # total data
        total_medicines += medicines
        total_forms += forms
        total_major_functions += major_functions
        total_functions += functions

    print("total medicines num: %d" % len(total_medicines))
    print("total forms num: %d" % len(total_forms))
    print("total major functions num: %d" % len(total_major_functions))
    print("total functions num: %d" % len(total_functions))

    total_medicines_set = set(total_medicines)
    total_forms_set = set(total_forms)
    total_major_functions_set = set(total_major_functions)
    total_functions_set = set(total_functions)
    print("medicines types num: %d" % len(total_medicines_set))
    print("forms types num: %d" % len(total_forms_set))
    print("major functions types num: %d" % len(total_major_functions_set))
    print("functions types num: %d" % len(total_functions_set))

    # statistics the medicines
    total_medicines_map = {}

    for med in total_medicines_set:
        if total_medicines.count(med) > 0:
            total_medicines_map[med] = total_medicines.count(med)

    # sort map
    total_medicines_sorted = sorted(total_medicines_map.items(), key=operator.itemgetter(1), reverse=True)
    print(total_medicines_sorted)

    # statistics the forms
    total_forms_map = {}
    for form in total_forms_set:
        if total_forms.count(form) > 0:
            total_forms_map[form] = total_forms.count(form)
    # sort forms map
    total_forms_sorted = sorted(total_forms_map.items(), key=operator.itemgetter(1), reverse=True)
    print(total_forms_sorted)

    # statistics the major functions
    total_major_functions_map = {}
    for mj in total_major_functions_set:
        if total_major_functions.count(mj) > 0:
            total_major_functions_map[mj] = total_major_functions.count(mj)
    # sort major function map
    total_major_functions_sorted = sorted(total_major_functions_map.items(), key=operator.itemgetter(1), reverse=True)
    print(total_major_functions_sorted)

    # statistics the functions
    total_functions_map = {}
    for func in total_functions_set:
        if total_functions.count(func) > 0:
            total_functions_map[func] = total_functions.count(func)
    # sorted functions map
    total_functions_sorted = sorted(total_functions_map.items(), key=operator.itemgetter(1), reverse=True)
    print(total_functions_sorted)

    # write to .csv file
    # 1. total medicines sorted
    with open(sheet_name+"_中药统计.csv", "w", encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(("中药名称", "数量"))
        for item in total_medicines_sorted:
            writer.writerow(item)

    # 2. total forms sorted
    with open(sheet_name+"_剂型统计.csv", "w", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(("剂型", "数量"))
        for item in total_forms_sorted:
            writer.writerow(item)

    # 3. total major functions sorted
    with open(sheet_name+"_主治统计.csv", "w", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(("主治", "数量"))
        for item in total_major_functions_sorted:
            writer.writerow(item)

    # 4. total functions sorted
    with open(sheet_name+"_功能统计.csv", "w", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(("功能", "数量"))
        for item in total_functions_sorted:
            writer.writerow(item)

    with open(sheet_name+"_patent_medicines.csv", "w", encoding="utf-8-sig") as f:
        for record in patent_medicines:
            for i in range(len(record)):
                if i == len(record) - 1:
                    f.write(record[i])
                else:
                    f.write(record[i] + ",")
            f.write("\n")

    print("sheet data processing finished!")


def main():
    path = "dataset/Copy of 桔梗专利汇总-6.28-2桔梗已替换.xlsx"

    # 1. Load excel file
    wb = load_workbook(path)

    # 2. Get all sheets of this excel file
    print("sheet names:", wb.sheetnames)

    # 3. Process the sheet
    # for name in wb.sheetnames:
    #     data_process(wb, name)

    data_process(wb, "Sheet1")


if __name__ == '__main__':
    main()
